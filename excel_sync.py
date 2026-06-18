# -*- coding: utf-8 -*-

import hashlib
import os
import threading
import unicodedata
from datetime import date, datetime, time as datetime_time
from io import BytesIO
from http.cookiejar import CookieJar
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
from urllib.request import HTTPCookieProcessor, Request, build_opener

from openpyxl import load_workbook

import database
from forms_parser import slug_operador, valor_ou_na


PLANILHA_URL = (os.environ.get("FORMS_PLANILHA_URL") or "").strip()
FONTE = (os.environ.get("FORMS_PLANILHA_FONTE") or "forms_ocorrencias").strip()
# O Forms precisa aparecer rapidamente na operacao; no Render o ciclo e fixo.
INTERVALO = 15
TIMEOUT = max(10, int(os.environ.get("FORMS_PLANILHA_TIMEOUT") or 30))

_thread = None
_thread_lock = threading.Lock()
_sync_lock = threading.Lock()
_parar = threading.Event()
_status_lock = threading.Lock()
_status = {
    "ativo": False,
    "fonte": FONTE,
    "intervalo_segundos": INTERVALO,
    "ultima_sincronizacao": None,
    "ultimo_erro": None,
    "ultimo_id": None,
    "importados_ultima_execucao": 0,
}


def _normalizar_cabecalho(valor):
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return " ".join(texto.strip().lower().replace("?", "").split())


def _texto(valor, padrao="N/A"):
    if valor is None:
        return padrao
    if isinstance(valor, float) and valor != valor:
        return padrao
    texto = str(valor).strip()
    return texto if texto else padrao


def _formatar_data(valor):
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")
    texto = _texto(valor, "")
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return texto


def _formatar_hora(valor):
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M:%S")
    if isinstance(valor, datetime_time):
        return valor.strftime("%H:%M:%S")
    texto = _texto(valor, "")
    if len(texto) == 5 and texto.count(":") == 1:
        return f"{texto}:00"
    return texto


def _combinar_data_hora(data_valor, hora_valor, fallback=None):
    data_texto = _formatar_data(data_valor)
    hora_texto = _formatar_hora(hora_valor)
    if data_texto and hora_texto:
        combinado = f"{data_texto} {hora_texto}"
        for formato in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(combinado, formato).isoformat(sep=" ")
            except ValueError:
                continue

    if isinstance(fallback, datetime):
        return fallback.replace(microsecond=0).isoformat(sep=" ")
    return database.agora_iso()


def _valor(linha, *nomes, padrao="N/A"):
    for nome in nomes:
        chave = _normalizar_cabecalho(nome)
        if chave in linha:
            valor = linha[chave]
            if valor is not None and str(valor).strip():
                return valor
    return padrao


def _id_linha(linha):
    valor = _valor(linha, "Id", padrao=None)
    if valor is None:
        return None
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None


def linha_para_payload(linha, fonte=FONTE):
    linha_id = _id_linha(linha)
    evento = _texto(_valor(linha, "EVENTO", padrao=""), "").upper()
    if linha_id is None:
        raise ValueError("Linha sem Id valido")
    if not evento:
        raise ValueError(f"Linha {linha_id} sem EVENTO")

    operador_nome = _texto(_valor(linha, "Nome"))
    operador_email = _texto(_valor(linha, "Email"))
    conclusao = _valor(linha, "Hora de conclusao", padrao=None)
    data_hora = _combinar_data_hora(
        _valor(linha, "DATA", padrao=None),
        _valor(linha, "HORA (hh:mm)", padrao=None),
        conclusao,
    )

    campos_adicionais = {
        "contrato_programacao": valor_ou_na(_valor(linha, "CLIENTE")),
        "programacao_carga": valor_ou_na(_valor(linha, "PROGRAMACAO DE CARGA")),
        "tempo_direcao_continua": valor_ou_na(_valor(linha, "TEMPO DE DIRECAO CONTINUA (hh:mm)")),
        "tempo_margem": valor_ou_na(_valor(linha, "TEMPO DE MARGEM (hh:mm)")),
        "tempo_estimado_destino": valor_ou_na(_valor(linha, "TEMPO ESTIMADO DO DESTINO (hh:mm)")),
        "motorista_sinal_fadiga": valor_ou_na(_valor(
            linha,
            "EM CASO DE PARADA PREVENTIVA, MOTORISTA APRESENTA SINAL DE FADIGA",
        )),
        "cobertura_celular": valor_ou_na(_valor(linha, "CELULAR ESTA COM AREA DE COBERTURA")),
        "motorista_atendeu": valor_ou_na(_valor(linha, "MOTORISTA ATENDEU A LIGACAO")),
        "mensagem_sirene": valor_ou_na(_valor(
            linha,
            "FOI ENVIADO MENSAGENS POR WHATSAPP, RASTREADOR E COMANDO DE SIRENE",
        )),
        "motorista_aceitou_parar": valor_ou_na(_valor(linha, "MOTORISTA ACEITOU PARAR")),
        "programacao_acionada": valor_ou_na(_valor(linha, "SETOR PROGRAMACAO FOI ACIONADO")),
        "programador": valor_ou_na(_valor(linha, "QUAL PROGRAMADOR NOS ATENDEU")),
        "evidencia": valor_ou_na(_valor(linha, "ANEXAR EVIDENCIA")),
    }

    return {
        "external_id": f"excel:{fonte}:{linha_id}",
        "excel_linha_id": linha_id,
        "evento": evento,
        "tipo": evento,
        "placa": _texto(_valor(linha, "PLACA")).upper(),
        "motorista": _texto(_valor(linha, "MOTORISTA")),
        "data_hora": data_hora,
        "operador_nome": operador_nome,
        "operador_email": operador_email,
        "operador": slug_operador(operador_nome, operador_email),
        "maquina": "EXCEL_FORMS",
        "origem": "EXCEL_FORMS",
        "campos_adicionais": campos_adicionais,
        "observacao_inicial": valor_ou_na(_valor(linha, "OBSERVACAO")),
    }


def _url_download_direto(url):
    partes = urlsplit(str(url or "").strip())
    if "sharepoint.com" not in partes.netloc.lower():
        return url

    parametros = parse_qsl(partes.query, keep_blank_values=True)
    if not any(chave.lower() == "download" for chave, _valor_query in parametros):
        parametros.append(("download", "1"))
    return urlunsplit((
        partes.scheme,
        partes.netloc,
        partes.path,
        urlencode(parametros),
        partes.fragment,
    ))


def _baixar_planilha(url):
    if os.path.exists(url):
        with open(url, "rb") as arquivo:
            return arquivo.read()

    url = _url_download_direto(url)

    request = Request(
        url,
        headers={
            "User-Agent": "MonfreTrack-ExcelSync/1.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    cookies = CookieJar()
    opener = build_opener(HTTPCookieProcessor(cookies))
    with opener.open(request, timeout=TIMEOUT) as resposta:
        conteudo = resposta.read()
        tipo_conteudo = resposta.headers.get("Content-Type", "desconhecido")
    if not conteudo:
        raise RuntimeError("A planilha retornou um arquivo vazio")
    if not conteudo.startswith(b"PK"):
        raise RuntimeError(
            f"O link nao retornou um arquivo XLSX (Content-Type: {tipo_conteudo})"
        )
    return conteudo


def _ler_linhas(conteudo):
    workbook = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        cabecalhos = next(rows, None)
        if not cabecalhos:
            return []

        chaves = [_normalizar_cabecalho(valor) for valor in cabecalhos]
        linhas = []
        for valores in rows:
            linha = {chaves[i]: valores[i] for i in range(min(len(chaves), len(valores)))}
            if _id_linha(linha) is not None:
                linhas.append(linha)
        return linhas
    finally:
        workbook.close()


def _atualizar_status(**campos):
    with _status_lock:
        _status.update(campos)


def status_sincronizador():
    with _status_lock:
        atual = dict(_status)
    banco = database.obter_estado_sincronizacao_excel(FONTE)
    if banco:
        atual.update({
            "ultimo_id": banco.get("ultimo_id"),
            "ultima_sincronizacao": banco.get("ultima_sincronizacao_em"),
            "ultimo_erro": banco.get("ultimo_erro"),
            "inicializado_em": banco.get("inicializado_em"),
        })
    atual["configurado"] = bool(PLANILHA_URL)
    return atual


def sincronizar_planilha(url=None, fonte=FONTE):
    url = (url or PLANILHA_URL).strip()
    if not url:
        return {"ok": False, "configurado": False, "erro": "FORMS_PLANILHA_URL nao configurada"}
    if os.environ.get("RENDER") and not database.USING_POSTGRES:
        erro = "DATABASE_URL nao configurado; sincronizacao Excel exige banco persistente no Render"
        database.registrar_erro_sincronizacao_excel(fonte, erro)
        _atualizar_status(ativo=False, ultimo_erro=erro)
        return {"ok": False, "erro": erro}

    if not _sync_lock.acquire(blocking=False):
        return {"ok": True, "ocupado": True, "importados": 0}

    try:
        conteudo = _baixar_planilha(url)
        linhas = sorted(_ler_linhas(conteudo), key=lambda item: _id_linha(item) or 0)
        maior_id = max((_id_linha(linha) or 0 for linha in linhas), default=0)
        estado = database.obter_estado_sincronizacao_excel(fonte)

        inicializado = not bool(estado)
        if inicializado:
            # Evita importar todo o historico, mas preserva a resposta que motivou
            # a ativacao do sincronizador.
            database.inicializar_sincronizacao_excel(fonte, max(0, maior_id - 1))
            estado = database.obter_estado_sincronizacao_excel(fonte)

        ultimo_id = int(estado.get("ultimo_id") or 0)
        novas = [linha for linha in linhas if (_id_linha(linha) or 0) > ultimo_id]
        recuperou_linha_corte = False
        if (
            not inicializado
            and ultimo_id > 0
            and not database.linha_excel_importada(fonte, ultimo_id)
        ):
            linha_corte = next(
                (linha for linha in linhas if _id_linha(linha) == ultimo_id),
                None,
            )
            if linha_corte is not None:
                novas.insert(0, linha_corte)
                recuperou_linha_corte = True
        importados = 0
        ignorados = 0

        for linha in novas:
            linha_id = _id_linha(linha)
            if database.linha_excel_importada(fonte, linha_id):
                database.atualizar_estado_sincronizacao_excel(fonte, linha_id)
                ultimo_id = max(ultimo_id, linha_id)
                continue

            try:
                payload = linha_para_payload(linha, fonte)
                ocorrencia_id = database.criar_ocorrencia_forms(payload)
                payload_hash = hashlib.sha256(repr(sorted(payload.items())).encode("utf-8")).hexdigest()
                database.registrar_linha_excel_importada(
                    fonte, linha_id, ocorrencia_id, payload_hash, "IMPORTADA", None
                )
                importados += 1
            except ValueError as exc:
                database.registrar_linha_excel_importada(
                    fonte, linha_id, None, None, "IGNORADA", str(exc)
                )
                ignorados += 1

            database.atualizar_estado_sincronizacao_excel(fonte, linha_id)
            ultimo_id = max(ultimo_id, linha_id)

        agora = database.agora_iso()
        database.concluir_sincronizacao_excel(fonte, ultimo_id)
        _atualizar_status(
            ativo=True,
            ultima_sincronizacao=agora,
            ultimo_erro=None,
            ultimo_id=ultimo_id,
            importados_ultima_execucao=importados,
        )
        return {
            "ok": True,
            "inicializado": inicializado,
            "ultimo_id": ultimo_id,
            "novas_linhas": len(novas),
            "importados": importados,
            "ignorados": ignorados,
            "recuperou_linha_corte": recuperou_linha_corte,
        }
    except Exception as exc:
        erro = f"{type(exc).__name__}: {exc}"
        database.registrar_erro_sincronizacao_excel(fonte, erro)
        _atualizar_status(ativo=False, ultimo_erro=erro, importados_ultima_execucao=0)
        return {"ok": False, "erro": erro}
    finally:
        _sync_lock.release()


def _loop_sincronizacao():
    _atualizar_status(ativo=True)
    while not _parar.is_set():
        resultado = sincronizar_planilha()
        if not resultado.get("ok"):
            print(f"Excel Sync: {resultado.get('erro')}")
        elif resultado.get("importados"):
            print(f"Excel Sync: {resultado['importados']} nova(s) ocorrencia(s) importada(s)")
        _parar.wait(INTERVALO)


def iniciar_sincronizador():
    global _thread
    if not PLANILHA_URL:
        _atualizar_status(ativo=False, ultimo_erro="FORMS_PLANILHA_URL nao configurada")
        return False

    with _thread_lock:
        if _thread and _thread.is_alive():
            return True
        _parar.clear()
        _thread = threading.Thread(
            target=_loop_sincronizacao,
            name="monfretrack-excel-sync",
            daemon=True,
        )
        _thread.start()
        return True


def parar_sincronizador():
    _parar.set()
